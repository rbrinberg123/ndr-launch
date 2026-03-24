import io
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', start_color='1F3864')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT   = Font(name='Arial', size=10)
ALT_FILL    = PatternFill('solid', start_color='EEF2F7')
NO_FILL     = PatternFill(fill_type=None)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left',   vertical='center')
THIN        = Border(bottom=Side(style='thin', color='D9D9D9'))

NUMERIC_COLS = {
    'Shares', 'Fund Shares', 'Passive or Index Shares',
    'Total Funds', 'Passive or Index Funds', 'Match Count',
    'EAUM ($mm)', 'AUM ($mm)', 'T/O %',
    'L12M', 'Total', '3rd Party', 'Rose & Co',
}
DATE_COLS  = {
    'Specifically with Co.', 'Anyone at Inst. with Co', 'Last Meeting', 'As of',
    'Last Mtg btwn Contact & Co', 'Last Mtg btwn firm & Co', 'Last Mtg. w/ Any Co',
}
SHRINK_COLS = {'Industry', 'Geo', 'Style', 'Mkt. Cap'}
SHRINK_ALIGN = Alignment(horizontal='left', vertical='center', shrink_to_fit=True)

# Reverse mcap map for display (translated values → human-readable)
MCAP_DISPLAY = {
    '****Micro': 'Micro',
    '***Small':  'Small',
    '**Mid':     'Mid',
    '*Large':    'Large',
    'Mega':      'Mega',
}


def _format_sheet(ws):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = Border(bottom=Side(style='medium', color='FFFFFF'))
    ws.row_dimensions[1].height = 20

    for row_idx in range(2, ws.max_row + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else NO_FILL
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = BODY_FONT
            cell.fill   = fill
            cell.border = THIN
            if header in NUMERIC_COLS:
                cell.alignment = CENTER
                if header == 'T/O %':
                    cell.number_format = '#,##0.0'
                elif header in ('EAUM ($mm)', 'AUM ($mm)'):
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0'
            elif header in DATE_COLS:
                cell.alignment    = CENTER
                cell.number_format = 'mm/dd/yyyy'
            elif header in SHRINK_COLS:
                cell.alignment = SHRINK_ALIGN
            else:
                cell.alignment = LEFT

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if header in SHRINK_COLS:
            ws.column_dimensions[col_letter].width = 27
        else:
            max_len = len(str(header)) if header else 8
            for row_idx in range(2, min(ws.max_row + 1, 300)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 42)

    ws.freeze_panes = 'A2'
    if ws.max_column > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"


def _write_summary_sheet(wb, results, company_name):
    """Insert a Summary sheet as the first sheet."""
    ws = wb.create_sheet('Summary', 0)

    NAVY        = PatternFill('solid', start_color='1F3864')
    NAVY_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    SEC_FILL    = PatternFill('solid', start_color='D6DCE4')
    SEC_FONT    = Font(name='Arial', bold=True, color='1F3864', size=9)
    LABEL_FONT  = Font(name='Arial', bold=True, size=10)
    VALUE_FONT  = Font(name='Arial', size=10)
    BORDER_THIN = Border(bottom=Side(style='thin', color='D9D9D9'))

    def fmt_list(vals):
        if not vals:
            return '(all / skip)'
        return ', '.join(sorted(str(v) for v in vals))

    def fmt_mcap(vals):
        if not vals:
            return '(all / skip)'
        return ', '.join(MCAP_DISPLAY.get(v, v) for v in sorted(vals))

    criteria = results.get('criteria', {})
    frames   = results.get('frames', {})

    hf_labels = {
        'separate':    'Move to HFs sheet',
        'include':     'Include in main results',
        'low_turnover':'Low-turnover only in main',
    }
    mtg_labels = {
        'include_all':  'Include all',
        'exclude_l12m': 'Exclude last 12 months',
        'exclude_l24m': 'Exclude last 24 months',
        'exclude_all':  'Exclude all prior meetings',
    }
    sh_labels = {
        'include_all': 'Include all',
        'exclude_all': 'Exclude all shareholders',
        'gt_001':      '> 0.01%',
        'gt_002':      '> 0.02%',
        'gt_003':      '> 0.03%',
        'gt_04':       '> 0.4%',
        'gt_05':       '> 0.5%',
    }
    routing_labels = {
        'virtual':           'Virtual',
        'investment_center': 'Investment Center',
        'cities':            'City',
        'state':             'State',
    }

    routing_mode  = results.get('routing_mode', 'virtual')
    virtual_scope = results.get('virtual_scope', 'both')
    city_sels     = results.get('city_selections') or []

    routing_display = routing_labels.get(routing_mode, routing_mode)
    if routing_mode == 'virtual' and virtual_scope != 'both':
        routing_display += f' ({virtual_scope.upper()})'
    elif city_sels:
        routing_display += ': ' + ', '.join(n for n, _ in city_sels)

    syms = results.get('subject_symbols', '')
    if isinstance(syms, list):
        syms = ', '.join(syms)

    eaum = results.get('eaum_min')

    # Build row data: (label, value, row_type)
    rows = [
        ('NDR Launch — Contact Filter Summary', None, 'title'),
        ('COMPANY',       None,                  'section'),
        ('Company',       company_name),
        ('Subject Ticker', syms or '—'),
        ('Generated',     str(date.today())),
        ('CDF CRITERIA',  None,                  'section'),
        ('Industry Focus',    fmt_list(criteria.get('industry'))),
        ('Investment Style',  fmt_list(criteria.get('style'))),
        ('Market Cap',        fmt_mcap(criteria.get('mcap'))),
        ('Geography',         fmt_list(criteria.get('geo'))),
        ('SETTINGS',      None,                  'section'),
        ('NDR Routing',          routing_display),
        ('Hedge Funds',          hf_labels.get(results.get('hf_treatment', 'separate'), '')),
        ('Meeting Exclusion',    mtg_labels.get(results.get('meeting_exclusion', 'include_all'), '')),
        ('Shareholder Exclusion',sh_labels.get(results.get('shareholder_exclusion', 'include_all'), '')),
        ('EAUM Minimum',         f'${eaum:,.0f}M' if eaum else '—'),
        ('RESULTS',       None,                  'section'),
        ('Source contacts',  results.get('total_source', 0)),
        ('Matched contacts', results.get('total_matched', 0)),
    ]

    sheet_order = ['Contacts', 'Virtual - NAM', 'Virtual - EUR', 'Virtual - Other',
                   'Too Small', 'Fixed Income', 'HFs', 'DNC', 'Check', 'Quant', 'Activist', 'Excluded']
    city_sels_names = {n for n, _ in city_sels} if city_sels else set()

    for name in city_sels_names:
        df = frames.get(name)
        if df is not None and len(df) > 0:
            rows.append((name, len(df)))

    for name in sheet_order:
        if name in city_sels_names:
            continue
        df = frames.get(name)
        if df is not None and len(df) > 0:
            rows.append((name, len(df)))

    # Write rows
    for i, row_data in enumerate(rows, start=1):
        row_type = row_data[2] if len(row_data) > 2 else 'data'
        label    = row_data[0]
        value    = row_data[1]

        if row_type == 'title':
            c = ws.cell(row=i, column=1, value=label)
            c.font      = NAVY_FONT
            c.fill      = NAVY
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'A{i}:B{i}')
            ws.row_dimensions[i].height = 24

        elif row_type == 'section':
            c = ws.cell(row=i, column=1, value=label)
            c.font      = SEC_FONT
            c.fill      = SEC_FILL
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'A{i}:B{i}')
            ws.row_dimensions[i].height = 16

        else:
            fill = ALT_FILL if i % 2 == 0 else NO_FILL
            val_str = str(value) if value is not None else '—'
            wrap    = len(val_str) > 60

            c1 = ws.cell(row=i, column=1, value=label)
            c1.font      = LABEL_FONT
            c1.fill      = fill
            c1.alignment = Alignment(horizontal='left', vertical='center')
            c1.border    = BORDER_THIN

            c2 = ws.cell(row=i, column=2, value=val_str)
            c2.font      = VALUE_FONT
            c2.fill      = fill
            c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=wrap)
            c2.border    = BORDER_THIN

            if wrap:
                ws.row_dimensions[i].height = max(30, min(len(val_str) // 3, 90))

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 82
    ws.freeze_panes = 'A2'


def generate_excel(results, company_name):
    frames          = results['frames']
    city_selections = results.get('city_selections') or []
    has_city        = results.get('has_city_routing', False)

    output = io.BytesIO()

    # Determine contact sheet write order
    sheet_order = []
    if has_city and city_selections:
        for tab_name, _ in city_selections:
            if tab_name in frames and frames[tab_name] is not None and len(frames[tab_name]) > 0:
                sheet_order.append(tab_name)
        for vname in ['Virtual - NAM', 'Virtual - EUR', 'Virtual - Other', 'Virtual']:
            if vname in frames and frames[vname] is not None and len(frames[vname]) > 0:
                sheet_order.append(vname)
    else:
        if 'Contacts' in frames and frames['Contacts'] is not None and len(frames['Contacts']) > 0:
            sheet_order.append('Contacts')

    for extra in ['Too Small', 'Fixed Income', 'HFs', 'DNC', 'Check', 'Quant', 'Activist', 'Excluded']:
        if extra in frames and frames[extra] is not None and len(frames[extra]) > 0:
            sheet_order.append(extra)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name in sheet_order:
            df = frames[sheet_name]
            if df is not None and len(df) > 0:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        if not sheet_order:
            pd.DataFrame(columns=['No results']).to_excel(writer, sheet_name='Contacts', index=False)

    output.seek(0)
    wb = load_workbook(output)

    # Format contact sheets
    for sheet_name in wb.sheetnames:
        _format_sheet(wb[sheet_name])

    # Insert Summary sheet first
    _write_summary_sheet(wb, results, company_name)

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.read()
